import cv2
import numpy as np
import mtcnn
from architecture import *
from train_v2 import normalize, l2_normalizer
from scipy.spatial.distance import cosine
from tensorflow.keras.models import load_model
import pickle
import threading
import pandas as pd
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles import PatternFill

class FaceRecognitionSystem:
    def __init__(self):
        self.confidence_t = 0.99
        self.recognition_t = 0.5
        self.required_size = (160, 160)
        self.face_detector = mtcnn.MTCNN()
        self.face_encoder = InceptionResNetV2()
        self.path_m = "facenet_keras_weights.h5"
        self.face_encoder.load_weights(self.path_m)
        self.encoding_dict = self.load_pickle('encodings\\encodings.pkl')

    def load_pickle(self, path):
        with open(path, 'rb') as f:
            encoding_dict = pickle.load(f)
        return encoding_dict

    def get_face(self, img, box):
        x1, y1, width, height = box
        x1, y1 = abs(x1), abs(y1)
        x2, y2 = x1 + width, y1 + height
        face = img[y1:y2, x1:x2]
        return face, (x1, y1), (x2, y2)

    def get_encode(self, face, size):
        face = normalize(face)
        face = cv2.resize(face, size)
        encode = self.face_encoder.predict(np.expand_dims(face, axis=0))[0]
        return encode

    def update_attendance(self, name, subject_attendance_file, datetime_str, buffer_lecture=False):
        lecture_date = datetime.now().strftime("%Y-%m-%d")
        unique_column_name = datetime_str
        if buffer_lecture:
            unique_column_name += "_Buffer"
        if not os.path.exists(subject_attendance_file):
            df = pd.DataFrame(columns=['Name', unique_column_name])
        else:
            df = pd.read_excel(subject_attendance_file)
            if unique_column_name not in df.columns:
                df[unique_column_name] = None
        if name not in df['Name'].values:
            new_entry = pd.DataFrame({'Name': [name], unique_column_name: 'Present'})
            df = pd.concat([df, new_entry], ignore_index=True)
            print(f"Attendance Marked For {name}")
        else:
            df.loc[df['Name'] == name, unique_column_name] = 'Present'
            print(f"{name} Already Present.")
        df[unique_column_name] = df[unique_column_name].fillna("ABSENT")
        total_ses = len(df.columns) - 3
        df['Presence Percentage'] = df.drop(columns=['Name']).apply(lambda row: (row == 'Present').sum(), axis=1) / total_ses * 100
        cols = [col for col in df.columns if col != 'Presence Percentage']
        cols.append('Presence Percentage')
        df = df[cols]
        df.to_excel(subject_attendance_file, index=False)
        workbook = openpyxl.load_workbook(subject_attendance_file)
        worksheet = workbook.active
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
        percentage_col = worksheet.max_column
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, percentage_col)
            if cell.value is not None:
                try:
                    percentage = float(cell.value)
                    if percentage < 50:
                        cell.fill = red_fill
                    elif 50 <= percentage < 75:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = green_fill
                except ValueError:
                    pass
        workbook.save(subject_attendance_file)

    def process_frame(self, frame, subject_name, datetime_str, buffer_lecture=False):
        img_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        results = self.face_detector.detect_faces(img_rgb)
        subject_attendance_file = os.path.join('subject_attendance', f'{subject_name}_attendance.xlsx')
        for res in results:
            if res['confidence'] < self.confidence_t:
                continue
            face, pt_1, pt_2 = self.get_face(img_rgb, res['box'])
            encode = self.get_encode(face, self.required_size)
            encode = l2_normalizer.transform(encode.reshape(1, -1))[0]
            name = 'unknown'
            distance = float("inf")
            for db_name, db_encode in self.encoding_dict.items():
                dist = cosine(db_encode, encode)
                if dist < self.recognition_t and dist < distance:
                    name = db_name
                    distance = dist
                    self.update_attendance(name, subject_attendance_file, datetime_str, buffer_lecture=buffer_lecture)
            if name == 'unknown':
                cv2.rectangle(frame, pt_1, pt_2, (0, 0, 255), 2)
                cv2.putText(frame, name, pt_1, cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 1)
            else:
                cv2.rectangle(frame, pt_1, pt_2, (0, 255, 0), 2)
                cv2.putText(frame, name + f'__{distance:.2f}', (pt_1[0], pt_1[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 1,
                            (0, 200, 200), 2)
        return frame

    def camera_thread(self, subject_name, datetime_str):
        cap = cv2.VideoCapture(0)
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                print("CAM NOT OPENED")
                break
            current_time = datetime.now().time()
            buffer_lecture = False
            if current_time >= datetime.strptime("17:00:00", "%H:%M:%S").time():
                buffer_lecture = True
            frame = self.process_frame(frame, subject_name, datetime_str, buffer_lecture)
            cv2.imshow('camera', frame)
            if cv2.waitKey(1) & 0xFF in [ord('q'), ord('Q')]:
                break
        cap.release()
        cv2.destroyAllWindows()

if __name__ == "__main__":
    os.makedirs('subject_attendance', exist_ok=True)
    frc = FaceRecognitionSystem()
    subject_name = input("Enter subject name: ")
    datetime_str = input("Enter date and time (YYYY-MM-DD HH:MM): ")
    if not subject_name:
        print("Subject name cannot be empty.")
    else:
        frc.camera_thread(subject_name, datetime_str)
